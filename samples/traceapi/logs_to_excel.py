import xlrd
import os
import openpyxl 
from openpyxl.chart import BarChart, Series, Reference

# #load the list of APIs hooked
# hooked_api_file = "Hooked_APIS.xlsx"
# wb = xlrd.open_workbook(hooked_api_file)
# sheet = wb.sheet_by_index(0)
# hooked_apis_ls = []
# number_of_rows = sheet.nrows
# for i in range(1, number_of_rows):
#     hooked_apis_ls.append(sheet.cell_value(i, 0))

# #dictionary of all hooked apis with the number of times they were called from ALL samples
# all_hooked_apis_dict = {}
# for hal in hooked_apis_ls:
#     all_hooked_apis_dict[hal] = 0

# #loop over all API log files (Dataset)
# dataset_dir = "ransomware_dataset_results"
# for filename in os.listdir(dataset_dir):
#     if filename.endswith(".txt"): 
#         try:
#             #open file
#             abs_path = os.path.join(dataset_dir, filename)
#             f = open(abs_path)

#             #dictionary that maps every called API to the number of times it has been called
#             hooked_apis_dict = {}
#             for hal in hooked_apis_ls:
#                 hooked_apis_dict[hal] = 0
            
#             #fill the dictionary
#             for l in f:
#                 for key in hooked_apis_dict.keys():
#                     if key in l:
#                         hooked_apis_dict[key] += 1
#                         all_hooked_apis_dict[key] += 1

#             #write the dictionary in excel file
#             wb = openpyxl.Workbook() 
#             sheet1 = wb.active 
#             sheet1.title = "Sheet1"
#             c1 = sheet1.cell(row = 1, column = 1) 
#             c2 = sheet1.cell(row = 1, column = 2) 
#             c1.value = "API"
#             c2.value = "nb_times"
#             row = 2
#             api_col = 1
#             counter_col = 2
#             for api, nb_times in hooked_apis_dict.items():
#                 sheet1.cell(row = row, column = api_col).value = api
#                 sheet1.cell(row = row, column = counter_col).value = nb_times
#                 row += 1

#             # draw the chart
#             chart1 = BarChart()
#             chart1.type = "col"
#             chart1.style = 10
#             chart1.title = "Bar Chart"
#             chart1.x_axis.title = 'list of APIs'
#             chart1.y_axis.title = 'number of times called'

#             data = Reference(sheet1, min_col=2, min_row=1, max_row=len(hooked_apis_ls), max_col=2)
#             cats = Reference(sheet1, min_col=1, min_row=2, max_row=len(hooked_apis_ls))
#             chart1.add_data(data, titles_from_data=True)
#             chart1.set_categories(cats)
#             chart1.shape = 4
#             chart1.height = 11.684
#             chart1.width = 34.8742
#             sheet1.add_chart(chart1, "D3")
            
#             wb.save(dataset_dir + '/' + filename[:-3] + 'xlsx')

#         except:
#             print("error in ", filename)


# #Now writing all_hooked_apis_dict into new excel sheet

# #write the dictionary into the excel file
# wb = openpyxl.Workbook() 
# sheet1 = wb.active 
# sheet1.title = "Sheet1"
# c1 = sheet1.cell(row = 1, column = 1) 
# c2 = sheet1.cell(row = 1, column = 2) 
# c1.value = "API"
# c2.value = "nb_times"
# row = 2
# api_col = 1
# counter_col = 2
# for api, nb_times in all_hooked_apis_dict.items():
#     sheet1.cell(row = row, column = api_col).value = api
#     sheet1.cell(row = row, column = counter_col).value = nb_times
#     row += 1

# # specify the shape
# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.x_axis.title = 'list of APIs'
# chart1.y_axis.title = 'number of times called from all samples'

# #assign the data
# data = Reference(sheet1, min_col=2, min_row=1, max_row=len(all_hooked_apis_dict), max_col=2)
# cats = Reference(sheet1, min_col=1, min_row=2, max_row=len(all_hooked_apis_dict))
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# chart1.height = 11.684
# chart1.width = 34.8742
# sheet1.add_chart(chart1, "D3")

# wb.save(dataset_dir + '/' + "All_samples.xlsx")

def sum_logs():
    #load the apis with their rank
    moni = {}
    f = open("mon_names.txt")
    for l in f:
        l = l.split("\t")
        moni[l[0].strip()] = float(l[1].strip())

    #load aliases
    aliases = []
    f = open("mon_alias.txt")
    for l in f:
        l = l.split("\t")
        aliases.append(l)

    

    dataset_dir = "ransomware_dataset_results"
    for filename in os.listdir(dataset_dir):
        
        #open logs file
        if filename.endswith(".txt") and filename == "rs1.txt": 
            try:
                abs_path = os.path.join(dataset_dir, filename)
                f = open(abs_path)
                sum = 0
                parsed_moni = {}
                for l in f:
                    # l = l.lower()

                    for m in moni:
                        # m = m.lower()

                        if m in l and m not in parsed_moni:

                            #check if any of its aliases has been called
                            flag = 1
                            for k in parsed_moni.keys():
                                # k = k.lower()

                                for a in aliases:
                                    # a = a.lower()

                                    #the called API, and the previous alias in the same subarray
                                    if k in a and m in a:
                                        flag = 0

                            if flag == 1:
                                sum += moni[m]
                                parsed_moni[m] = moni[m]
                print(filename, sum)
                [print (k, v) for k, v in parsed_moni.items()]
            except Exception as e:
                print(e)

sum_logs()